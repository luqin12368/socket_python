#依赖python-nmap,openpyxl包
import nmap
import openpyxl
import time
import IPy
from multiprocessing import Pool as ThreadPool

# 同目录下创建一个input.txt,放入ip地址列(从xlsx中直接粘贴即可)
expath=''
inputFile='input.txt'
th=20

port='60008'
#默认命令
arg='-Pn -T4'
#输出文件名
outPutFile = '123.txt'


def readfile(path):
    # 返回url列表
    file = open(path, 'r',encoding='utf-8')
    urlList=[]
    for eachline in file.readlines():
        #获得地址段
        url=eachline.rsplit()[0]
        urlList.append(url)


    file.close()
    # 造infoList
    infoList=[]
    for url in urlList:
        info={'hostname':url,'port':port,'arg':arg}
        infoList.append(info)

    return infoList

def scan(info):

    hostName=info['hostname']
    port=info['port']
    print(hostName+'扫描开始')

    scanner=nmap.PortScanner()
    scanner.scan(hosts=hostName,ports=port,arguments=arg)

    # print(res)
    out=open(outPutFile,'a+',encoding='utf-8')
    # out.write()
    state=scanner[hostName].state()
    tcp_ports=scanner[hostName].all_tcp()
    print(hostName+' '+state)
    for ports in tcp_ports:
        if scanner[hostName]['tcp'][ports]['state'] =='open' :
            st=hostName+' '+str(ports)+' '+scanner[hostName]['tcp'][ports]['name']+' '+scanner[hostName]['tcp'][ports]['state']+'\n'
            out.write(st)
            print(st)

    # print(hostName)
    out.close()


def makeEx():
    print('----------------开始写入-------------------')
    wb=openpyxl.load_workbook(expath)
    sheet=wb.active
    c=open(outPutFile,'r',encoding='utf-8')
    i=1
    for line in c.readlines():
        linelist=line.split()
        ip=linelist[0]
        port=linelist[1]
        name=linelist[2]
        status=linelist[3]
        for j in range(4):
            if j==3:
                n=linelist[j]
                n=n.rsplit()
                n=n[0]
                sheet.cell(row=i, column=j + 1, value=n)
            else:
                sheet.cell(row=i,column=j+1,value=linelist[j])
        i=i+1
    c.close()
    wb.save(expath)
    print('----------------写入完成-------------------')


if __name__ == '__main__':
    print('----------------扫描开始-------------------')
    start=time.time()
    pool = ThreadPool(th)
    infolist=readfile(inputFile)
    pool.map(scan,infolist)
    pool.close()
    pool.join()
    finish=time.time()
    print('----------------扫描完成-------------------')
    t=finish-start
    print('用时 %f'%t)
    with open(outPutFile, 'a+', encoding='utf-8') as out:
        out.write('\n用时 %f'%t)
